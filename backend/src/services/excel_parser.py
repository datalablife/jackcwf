"""
Excel file parser service for extracting and processing Excel data.

Provides functionality to parse .xlsx and .xls files, list sheets,
extract column information, infer data types, and generate previews.
"""

from typing import List, Dict, Any, Optional
from pathlib import Path


class ExcelParserService:
    """
    Service for parsing Excel files (.xlsx and .xls formats).

    Provides methods to extract structure, list sheets, infer types, and generate previews.
    """

    # Data type inference patterns
    EMPTY_VALUE_INDICATORS = {"", "null", "none", "na", "n/a", "unknown", "nan", "inf"}

    @staticmethod
    def parse_excel(
        file_path: str,
        sheet_name: Optional[str] = None,
        sheet_index: int = 0
    ) -> Dict[str, Any]:
        """
        Parse an Excel file and extract its structure.

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to parse (default: first sheet)
            sheet_index: Index of the sheet (default: 0)

        Returns:
            Dictionary with file structure information

        Raises:
            Exception: If file cannot be parsed

        Example:
            >>> result = ExcelParserService.parse_excel("data.xlsx")
            >>> result["row_count"]
            5000
        """
        try:
            file_ext = Path(file_path).suffix.lower()

            if file_ext == ".xlsx":
                return ExcelParserService._parse_xlsx(file_path, sheet_name, sheet_index)
            elif file_ext == ".xls":
                return ExcelParserService._parse_xls(file_path, sheet_name, sheet_index)
            else:
                raise Exception(f"Unsupported Excel format: {file_ext}")

        except Exception as e:
            raise Exception(f"Failed to parse Excel: {str(e)}")

    @staticmethod
    def list_sheets(file_path: str) -> List[str]:
        """
        List all sheet names in an Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            List of sheet names

        Example:
            >>> ExcelParserService.list_sheets("data.xlsx")
            ["Sheet1", "Sheet2", "Sales"]
        """
        try:
            file_ext = Path(file_path).suffix.lower()

            if file_ext == ".xlsx":
                import openpyxl

                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheets = workbook.sheetnames
                workbook.close()
                return sheets

            elif file_ext == ".xls":
                import xlrd

                workbook = xlrd.open_workbook(file_path)
                return workbook.sheet_names()

            else:
                raise Exception(f"Unsupported Excel format: {file_ext}")

        except Exception as e:
            raise Exception(f"Failed to list sheets: {str(e)}")

    @staticmethod
    def get_sheet_data(
        file_path: str,
        sheet_name: Optional[str] = None,
        sheet_index: int = 0
    ) -> Dict[str, Any]:
        """
        Get all data from a specific sheet.

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet (default: first sheet)
            sheet_index: Index of the sheet (default: 0)

        Returns:
            Dictionary with sheet data

        Example:
            >>> data = ExcelParserService.get_sheet_data("data.xlsx", sheet_name="Sales")
            >>> len(data["rows"])
            1000
        """
        try:
            file_ext = Path(file_path).suffix.lower()

            if file_ext == ".xlsx":
                return ExcelParserService._get_xlsx_sheet_data(file_path, sheet_name, sheet_index)
            elif file_ext == ".xls":
                return ExcelParserService._get_xls_sheet_data(file_path, sheet_name, sheet_index)
            else:
                raise Exception(f"Unsupported Excel format: {file_ext}")

        except Exception as e:
            raise Exception(f"Failed to get sheet data: {str(e)}")

    @staticmethod
    def get_column_names(
        file_path: str,
        sheet_name: Optional[str] = None,
        sheet_index: int = 0
    ) -> List[str]:
        """
        Get column names from an Excel sheet.

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet (default: first sheet)
            sheet_index: Index of the sheet (default: 0)

        Returns:
            List of column names

        Example:
            >>> ExcelParserService.get_column_names("data.xlsx")
            ["id", "name", "email", "age"]
        """
        try:
            file_ext = Path(file_path).suffix.lower()

            if file_ext == ".xlsx":
                import openpyxl

                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheet = ExcelParserService._get_xlsx_sheet(workbook, sheet_name, sheet_index)

                headers = []
                for cell in sheet[1]:
                    if cell.value:
                        headers.append(str(cell.value))
                    else:
                        headers.append(f"Column_{len(headers) + 1}")

                workbook.close()
                return headers

            elif file_ext == ".xls":
                import xlrd

                workbook = xlrd.open_workbook(file_path)
                sheet = ExcelParserService._get_xls_sheet(workbook, sheet_name, sheet_index)
                return [str(cell.value) for cell in sheet.row(0)]

            else:
                raise Exception(f"Unsupported Excel format: {file_ext}")

        except Exception as e:
            raise Exception(f"Failed to get column names: {str(e)}")

    @staticmethod
    def get_data_types(
        file_path: str,
        sheet_name: Optional[str] = None,
        sheet_index: int = 0
    ) -> List[str]:
        """
        Infer data types from an Excel sheet.

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet (default: first sheet)
            sheet_index: Index of the sheet (default: 0)

        Returns:
            List of inferred data type strings

        Example:
            >>> ExcelParserService.get_data_types("data.xlsx")
            ["integer", "string", "string", "integer"]
        """
        try:
            sheet_data = ExcelParserService.get_sheet_data(file_path, sheet_name, sheet_index)
            column_names = sheet_data.get("column_names", [])
            rows = sheet_data.get("rows", [])

            data_types = []
            for col_name in column_names:
                col_type = ExcelParserService._infer_column_type(rows, col_name)
                data_types.append(col_type)

            return data_types

        except Exception as e:
            raise Exception(f"Failed to infer data types: {str(e)}")

    @staticmethod
    def get_preview(
        file_path: str,
        max_rows: int = 20,
        sheet_name: Optional[str] = None,
        sheet_index: int = 0
    ) -> Dict[str, Any]:
        """
        Get a preview of Excel sheet data.

        Args:
            file_path: Path to the Excel file
            max_rows: Maximum number of rows to return (default: 20)
            sheet_name: Name of the sheet (default: first sheet)
            sheet_index: Index of the sheet (default: 0)

        Returns:
            Dictionary with preview data

        Example:
            >>> preview = ExcelParserService.get_preview("data.xlsx")
            >>> len(preview["rows"])
            20
        """
        try:
            sheet_data = ExcelParserService.get_sheet_data(file_path, sheet_name, sheet_index)
            column_names = sheet_data.get("column_names", [])
            rows = sheet_data.get("rows", [])[:max_rows]

            return {
                "column_names": column_names,
                "rows": rows,
                "displayed_rows": len(rows),
                "max_rows": max_rows,
            }

        except Exception as e:
            raise Exception(f"Failed to get preview: {str(e)}")

    # Private helper methods for xlsx parsing
    @staticmethod
    def _parse_xlsx(
        file_path: str,
        sheet_name: Optional[str] = None,
        sheet_index: int = 0
    ) -> Dict[str, Any]:
        """Parse .xlsx file using openpyxl."""
        import openpyxl

        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = ExcelParserService._get_xlsx_sheet(workbook, sheet_name, sheet_index)

        # Extract headers and data
        headers = []
        rows = []
        row_count = 0

        for row_idx, row_cells in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx == 0:
                # Header row
                for idx, cell_value in enumerate(row_cells):
                    if cell_value:
                        headers.append(str(cell_value))
                    else:
                        headers.append(f"Column_{idx + 1}")
            else:
                # Data rows
                if any(cell_value is not None for cell_value in row_cells):
                    row_count += 1
                    row_dict = {}
                    for col_idx, cell_value in enumerate(row_cells):
                        if col_idx < len(headers):
                            row_dict[headers[col_idx]] = cell_value

                    rows.append(row_dict)

        # Infer data types
        data_types = []
        for col_name in headers:
            col_type = ExcelParserService._infer_column_type(rows, col_name)
            data_types.append(col_type)

        workbook.close()

        return {
            "row_count": row_count,
            "column_names": headers,
            "data_types": data_types,
            "columns_info": [
                {
                    "name": col_name,
                    "data_type": data_type,
                }
                for col_name, data_type in zip(headers, data_types)
            ],
        }

    @staticmethod
    def _parse_xls(
        file_path: str,
        sheet_name: Optional[str] = None,
        sheet_index: int = 0
    ) -> Dict[str, Any]:
        """Parse .xls file using xlrd."""
        import xlrd

        workbook = xlrd.open_workbook(file_path)
        sheet = ExcelParserService._get_xls_sheet(workbook, sheet_name, sheet_index)

        # Extract headers and data
        headers = []
        rows = []

        for col_idx, cell_value in enumerate(sheet.row(0)):
            if cell_value.value:
                headers.append(str(cell_value.value))
            else:
                headers.append(f"Column_{col_idx + 1}")

        # Extract data rows
        for row_idx in range(1, sheet.nrows):
            row_dict = {}
            row_values = sheet.row(row_idx)

            for col_idx, cell in enumerate(row_values):
                if col_idx < len(headers):
                    row_dict[headers[col_idx]] = cell.value

            rows.append(row_dict)

        # Infer data types
        data_types = []
        for col_name in headers:
            col_type = ExcelParserService._infer_column_type(rows, col_name)
            data_types.append(col_type)

        return {
            "row_count": len(rows),
            "column_names": headers,
            "data_types": data_types,
            "columns_info": [
                {
                    "name": col_name,
                    "data_type": data_type,
                }
                for col_name, data_type in zip(headers, data_types)
            ],
        }

    @staticmethod
    def _get_xlsx_sheet(workbook, sheet_name: Optional[str], sheet_index: int):
        """Get sheet object from xlsx workbook."""
        if sheet_name:
            return workbook[sheet_name]
        else:
            return workbook.worksheets[sheet_index]

    @staticmethod
    def _get_xls_sheet(workbook, sheet_name: Optional[str], sheet_index: int):
        """Get sheet object from xls workbook."""
        if sheet_name:
            return workbook.sheet_by_name(sheet_name)
        else:
            return workbook.sheet_by_index(sheet_index)

    @staticmethod
    def _get_xlsx_sheet_data(
        file_path: str,
        sheet_name: Optional[str],
        sheet_index: int
    ) -> Dict[str, Any]:
        """Get all data from xlsx sheet."""
        import openpyxl

        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = ExcelParserService._get_xlsx_sheet(workbook, sheet_name, sheet_index)

        headers = []
        rows = []

        for row_idx, row_cells in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx == 0:
                headers = [str(v) if v else f"Column_{i + 1}" for i, v in enumerate(row_cells)]
            else:
                if any(v is not None for v in row_cells):
                    row_dict = {headers[i]: v for i, v in enumerate(row_cells) if i < len(headers)}
                    rows.append(row_dict)

        workbook.close()

        return {
            "column_names": headers,
            "rows": rows,
        }

    @staticmethod
    def _get_xls_sheet_data(
        file_path: str,
        sheet_name: Optional[str],
        sheet_index: int
    ) -> Dict[str, Any]:
        """Get all data from xls sheet."""
        import xlrd

        workbook = xlrd.open_workbook(file_path)
        sheet = ExcelParserService._get_xls_sheet(workbook, sheet_name, sheet_index)

        headers = [str(cell.value) for cell in sheet.row(0)]
        rows = []

        for row_idx in range(1, sheet.nrows):
            row_values = sheet.row(row_idx)
            row_dict = {headers[i]: cell.value for i, cell in enumerate(row_values) if i < len(headers)}
            rows.append(row_dict)

        return {
            "column_names": headers,
            "rows": rows,
        }

    @staticmethod
    def _infer_column_type(rows: List[Dict[str, Any]], column_name: str) -> str:
        """Infer the data type of a single column."""
        if not rows:
            return "string"

        # Sample first 100 non-empty values
        sample_values = []
        for row in rows[:100]:
            value = row.get(column_name, "")
            if value and str(value).lower() not in ExcelParserService.EMPTY_VALUE_INDICATORS:
                sample_values.append(str(value).strip())

        if not sample_values:
            return "string"

        # Try to infer type
        all_int = True
        all_float = True
        all_bool = True

        for val in sample_values:
            # Check boolean
            if val.lower() not in {"true", "false", "yes", "no", "1", "0"}:
                all_bool = False

            # Check integer
            try:
                int(val)
            except ValueError:
                all_int = False

            # Check float
            try:
                float(val)
            except ValueError:
                all_float = False

        if all_bool:
            return "boolean"
        elif all_int:
            return "integer"
        elif all_float:
            return "float"
        else:
            return "string"
