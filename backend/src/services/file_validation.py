"""
File validation service for uploaded files.

Validates file types, sizes, and formats before processing.
"""

import os
from typing import Optional
from pathlib import Path


# Configuration constants
MAX_FILE_SIZE = 500 * 1024 * 1024  # 500 MB in bytes
ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls", ".json", ".jsonl"}
ALLOWED_MIME_TYPES = {
    "text/csv",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel",  # .xls
    "application/json",
}


class FileValidationError(Exception):
    """Exception raised when file validation fails."""

    def __init__(self, message: str, error_code: str = "VALIDATION_ERROR"):
        self.message = message
        self.error_code = error_code
        super().__init__(self.message)


class FileValidationService:
    """
    Service for validating uploaded files.

    Provides methods to validate file size, type, format, and content.
    """

    @staticmethod
    def validate_file(
        file_path: str,
        file_size: int,
        filename: str,
    ) -> bool:
        """
        Validate a file for upload.

        Checks:
        - File extension is in allowed list
        - File size doesn't exceed limit
        - File exists

        Args:
            file_path: Path to the file
            file_size: Size of the file in bytes
            filename: Original filename

        Returns:
            True if file is valid

        Raises:
            FileValidationError: If validation fails

        Example:
            >>> FileValidationService.validate_file("/tmp/data.csv", 1024000, "data.csv")
            True
        """
        # Check file extension
        file_ext = Path(filename).suffix.lower()
        if file_ext not in ALLOWED_EXTENSIONS:
            allowed = ", ".join(ALLOWED_EXTENSIONS)
            raise FileValidationError(
                f"File extension '{file_ext}' not allowed. Allowed types: {allowed}",
                "INVALID_FILE_TYPE",
            )

        # Check file size
        if file_size > MAX_FILE_SIZE:
            max_mb = MAX_FILE_SIZE // (1024 * 1024)
            raise FileValidationError(
                f"File size {file_size / (1024 * 1024):.1f}MB exceeds maximum {max_mb}MB",
                "FILE_TOO_LARGE",
            )

        # Check file exists (if path is provided)
        if file_path and not os.path.exists(file_path):
            raise FileValidationError(f"File not found at {file_path}", "FILE_NOT_FOUND")

        return True

    @staticmethod
    def validate_csv(file_path: str) -> bool:
        """
        Validate a CSV file format.

        Checks:
        - File can be read as CSV
        - File has headers
        - File has at least one row of data

        Args:
            file_path: Path to the CSV file

        Returns:
            True if file is valid CSV

        Raises:
            FileValidationError: If file is not valid CSV

        Example:
            >>> FileValidationService.validate_csv("/tmp/data.csv")
            True
        """
        try:
            import csv

            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                headers = next(reader, None)

                if not headers:
                    raise FileValidationError("CSV file is empty or has no headers", "INVALID_CSV_FORMAT")

                # Try to read at least one data row
                try:
                    next(reader)
                except StopIteration:
                    raise FileValidationError("CSV file has headers but no data rows", "INVALID_CSV_FORMAT")

            return True

        except UnicodeDecodeError:
            raise FileValidationError("CSV file is not UTF-8 encoded", "ENCODING_ERROR")
        except Exception as e:
            raise FileValidationError(f"Failed to validate CSV file: {str(e)}", "CSV_VALIDATION_ERROR")

    @staticmethod
    def validate_excel(file_path: str) -> bool:
        """
        Validate an Excel file format.

        Checks:
        - File is valid Excel format
        - File has at least one sheet
        - First sheet has headers and data

        Args:
            file_path: Path to the Excel file

        Returns:
            True if file is valid Excel

        Raises:
            FileValidationError: If file is not valid Excel

        Example:
            >>> FileValidationService.validate_excel("/tmp/data.xlsx")
            True
        """
        try:
            import openpyxl
            import xlrd

            file_ext = Path(file_path).suffix.lower()

            if file_ext == ".xlsx":
                try:
                    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    if not workbook.sheetnames:
                        raise FileValidationError("Excel file has no sheets", "INVALID_EXCEL_FORMAT")

                    sheet = workbook.active
                    if not sheet or sheet.max_row < 1:
                        raise FileValidationError("Excel sheet is empty", "INVALID_EXCEL_FORMAT")

                    workbook.close()
                except FileValidationError:
                    raise
                except Exception as e:
                    raise FileValidationError(f"Failed to read .xlsx file: {str(e)}", "INVALID_EXCEL_FORMAT")

            elif file_ext == ".xls":
                try:
                    workbook = xlrd.open_workbook(file_path)
                    if workbook.nsheets == 0:
                        raise FileValidationError("Excel file has no sheets", "INVALID_EXCEL_FORMAT")

                    sheet = workbook.sheet_by_index(0)
                    if sheet.nrows < 1:
                        raise FileValidationError("Excel sheet is empty", "INVALID_EXCEL_FORMAT")
                except FileValidationError:
                    raise
                except Exception as e:
                    raise FileValidationError(f"Failed to read .xls file: {str(e)}", "INVALID_EXCEL_FORMAT")

            return True

        except FileValidationError:
            raise
        except Exception as e:
            raise FileValidationError(f"Excel validation error: {str(e)}", "EXCEL_VALIDATION_ERROR")

    @staticmethod
    def validate_json(file_path: str) -> bool:
        """
        Validate a JSON file format.

        Checks:
        - File is valid JSON
        - File is not empty

        Args:
            file_path: Path to the JSON file

        Returns:
            True if file is valid JSON

        Raises:
            FileValidationError: If file is not valid JSON

        Example:
            >>> FileValidationService.validate_json("/tmp/data.json")
            True
        """
        try:
            import json

            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)

                if not data:
                    raise FileValidationError("JSON file is empty", "INVALID_JSON_FORMAT")

            return True

        except json.JSONDecodeError as e:
            raise FileValidationError(f"Invalid JSON format: {str(e)}", "INVALID_JSON_FORMAT")
        except UnicodeDecodeError:
            raise FileValidationError("JSON file is not UTF-8 encoded", "ENCODING_ERROR")
        except Exception as e:
            raise FileValidationError(f"JSON validation error: {str(e)}", "JSON_VALIDATION_ERROR")
